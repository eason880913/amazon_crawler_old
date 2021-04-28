from openpyxl import load_workbook
import time
import re
from bs4 import BeautifulSoup
import requests
import math
from fake_headers import Headers
from tqdm import tqdm

'''
待處理：
'''

inputfilename = 'AMAZON PHONE_20210418.xlsx'

def get_pages(url):
    headers = Headers(os="mac", headers=True).generate()
    ans = requests.get(url,headers=headers)
    soup = BeautifulSoup(ans.content,'lxml')
    count_comm = soup.select('[data-hook="cr-filter-info-review-rating-count"]')[0].text.strip().replace(',','')
    count_comm = int(re.findall('\d+',count_comm)[1])
    pages = math.ceil(count_comm/10)
    return pages

def get_rating(comm):
    try:
        if len(comm.select('[data-hook="review-star-rating"]')) > 0:
            rating = comm.select('[data-hook="review-star-rating"]')[0].text
        elif len(comm.select('[data-hook="cmps-review-star-rating"]')) > 0:
            rating = comm.select('[data-hook="cmps-review-star-rating"]')[0].text              
    except Exception as e:
        print(e)
        rating = 'error1'
    try:
        if len(re.findall('\d\\.0',rating)) > 0 :
            rating = re.findall('\d\\.0',rating)[0].replace('.0','').strip()
        elif len(re.findall('\d\\,0',rating)) > 0 :
            rating = re.findall('\d\\,0',rating)[0].replace(',0','').strip()
    except Exception as e:
        print(e)
        rating = 'error2'
    return rating

def main(inputfilename):
    wb = load_workbook(inputfilename)
    sheet_names = wb.get_sheet_names()#方法得到工作簿的所有工作表
    ws = wb[sheet_names[0]]
    all_dict = {"USA":12, "Canada":14, "Brazil":16, "turkey":18, "Australia":20, "UK":22, "Italy":24, "Nederland":26}
    for con in all_dict:
        outputfilename = con
        for i in range(2,ws.max_row+1): # 1爲編號
            #取得編號 產品名 網址
            number = (ws.cell(row=i, column=1).value)
            url = str((ws.cell(row=i, column=all_dict[con]).value))
            url = re.sub('dp','product-reviews',url)+'&sortBy=recent&pageNumber=1'
            # url = 'https://www.amazon.com/-/zh_TW/Moto-Power-Unlocked-International-Camera/dp/B087CBMKSC/ref=sr_1_1?dchild=1&keywords=Motorola%2BMoto%2BG8%2BPower&qid=1589849495&sr=8-1&th=1'    

            if url == None or 'http' not in url:
                continue
            try:
                pages = get_pages(url)
            except:
                continue

            for j in range(1, 1+int(pages)):
                url = re.sub('pageNumber=.*',f'pageNumber={j}',url)
                headers = Headers(os="mac", headers=True).generate()
                ans = requests.get(url,headers=headers)
                soup = BeautifulSoup(ans.content,'lxml')
                comments = soup.select('[data-hook="review"]')

                for comm in comments:
                    name = comm.select('[class="a-profile-name"]')[0].text.strip().replace(',',';')
                    rating = get_rating(comm)
                    title = comm.select('[data-hook="review-title"]')[0].text.strip().replace(',',';')
                    rec_time = comm.select('[data-hook="review-date"]')[0].text.strip().replace(',',';')
                    body = comm.select('[data-hook="review-body"]')[0].text.strip().replace('\n','').replace(',',';')
                    body = re.sub('Your browser does not support HTML5 video.\n\n\n','',body)
                    if len(comm.select('[data-hook="helpful-vote-statement"]')) > 0:
                        helpful = comm.select('[data-hook="helpful-vote-statement"]')[0].text
                        helpful = re.findall('\d+',rating)[0].strip()
                    else:
                        helpful = '0'

                    with open (f'phonecomment.csv','a', encoding= 'utf-8') as f:
                        f.write(f"{outputfilename}, {number}, {name}, {rating}, {title}, {rec_time}, {body}, {helpful}\n")

if __name__ == "__main__":
    main(inputfilename)