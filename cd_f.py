from openpyxl import load_workbook
import time
import re
import time
from selenium import webdriver
from bs4 import BeautifulSoup
import requests


driver_path = '/Users/eason880913/Desktop/work/fb_crawler/Internet-Observation-Station/chromedriver'
inputfilename = '20210418AMAZON-CD.xlsx'

def init_driver(driver_path):
    chrome_options = webdriver.ChromeOptions()
    prefs = {"profile.default_content_setting_values.notifications" : 2}
    #chrome_options.add_argument('--headless')
    chrome_options.add_argument('--max_old_space_size')    
    chrome_options.add_experimental_option("prefs",prefs) # turn of notification window\
    driver = webdriver.Chrome(driver_path,options=chrome_options)
    return driver 

def get_rating(soup):
    #rating
    try:
        rating = soup.select('[class="a-icon-alt"]')[0].text.split()[0]
        rating = re.sub('\n','',rating)
        rating = re.sub('\\,',';',rating)
        rating = re.sub('5つ星のうち','',rating)
    except:
        rating = ''
    return rating

def get_ranking(soup):
    #ranking
    try:
        ranking = soup.select('[class="a-unordered-list a-nostyle a-vertical a-spacing-none detail-bullet-list"]')[1].text
        ranking = re.sub('\n','',ranking)
        ranking = re.sub('\\,',';',ranking)
    except:
        ranking = ''
    return ranking

def get_price(soup, driver):
    #price
    try:
        price = soup.select('[id="price_inside_buybox"]')[0].text
        price = re.sub('\n','',price)
        price = re.sub('\\,',';',price)
        price = price.strip()
    except:
        price = ''
    if price == '': 
        try:
            driver.find_elements_by_css_selector('[id="buybox-see-all-buying-choices"] [class="a-button-inner"] [role="button"]')[0].click()
            time.sleep(3)
            soup1 = BeautifulSoup(driver.page_source,'lxml')
            price_list = []
            for j in range(len(soup1.select('[id="aod-offer-heading"] h5'))):
                res1 = soup1.select('[id="aod-offer-heading"] h5')[j].text.strip()
                if res1 == "新品" or res1 == "全新品" or res1 == "New" or res1 == "Neu" or res1 == "Nuevo" or res1 == "Nuovo" or res1 == "Novo" or res1 == "Nieuw":
                    symbol = soup1.select('[class="a-price"] [aria-hidden="true"] [class="a-price-symbol"]')[j].text
                    # print(symbol)
                    try:
                        fraction = soup1.select('[class="a-price"] [aria-hidden="true"] [class="a-price-fraction"]')[j].text
                    except:
                        fraction = ''
                    whole = re.sub('"|,|\.','',soup1.select('[class="a-price"] [aria-hidden="true"] [class="a-price-whole"]')[j].text)
                    price_list.append(float(str(whole)+'.'+str(fraction)))
                # print(price_list)
            price = symbol+str(min(price_list))
            price = re.sub('\n','',price)
            price = re.sub('\\,',';',price)
        except:   
            price = '' 
    if price == '': 
        try:
            price = soup.select('[id="newBuyBoxPrice"]')[0].text
            price = re.sub('\n','',price)
            price = re.sub('\\,',';',price)
            price = price.strip()
        except:
            price = ''
    return price 

def main(inputfilename, driver):
    #爬取時間
    rec_time = time.strftime("%Y-%m-%d", time.localtime())    
    #open workbook
    wb = load_workbook(inputfilename)
    sheet_names = wb.get_sheet_names()#方法得到工作簿的所有工作表
    ws = wb[sheet_names[0]]
    all_dict = {"USA":9, "Japan":12, "Germany":15, "Australia":18, "Canada":21, "Mexico":24, "France":27, "Spain":30, "Italy":33, "Brasil":36, "UK":39, "Nederland":42, "Singapor":45}
    
    for con in all_dict:
        outputfilename = con
        for i in range(2,ws.max_row+1): # 1爲編號
            #取得編號 產品名 網址
            number = (ws.cell(row=i, column=1).value)
            name = (ws.cell(row=i, column=2).value)
            name = re.sub('\\,',';',name)
            url = str((ws.cell(row=i, column=all_dict[con]).value))

            if url == None or 'http' not in url:
                continue
            
            driver.get(url)
            soup = BeautifulSoup(driver.page_source,'lxml')
            # get rating
            rating = get_rating(soup)

            # get ranking
            ranking = get_ranking(soup)
            
            #get price 
            price = get_price(soup, driver)

            with open (f'cd_{outputfilename}.csv','a', encoding= 'utf-8') as f:
                f.write(f"{number},{name},{rec_time},{rating},{ranking},{url},{price}\n")

if __name__ == "__main__":
    driver = init_driver(driver_path)
    main(inputfilename, driver)
    driver.close()